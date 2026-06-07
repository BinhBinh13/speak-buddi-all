# speak-buddi-be/services/export_builders.py
# ─── Sinh file Excel/PDF cho báo cáo admin (S11.3) ───────────────────────────

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from services.pdf_fonts import FONT_BOLD, FONT_REGULAR, ensure_pdf_fonts

REPORT_LABELS = {
    "revenue": "Doanh thu",
    "users": "Người dùng",
    "learning": "Học tập & Kiểm tra",
    "ai_usage": "Sử dụng AI",
}

EMPTY_MSG = "Chưa có dữ liệu trong khoảng thời gian đã chọn."


def _format_vnd(value: int | float | None) -> str:
    if value is None:
        return "0 ₫"
    return f"{int(value):,} ₫".replace(",", ".")


def build_xlsx(report_type: str, payload: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tóm tắt"
    bold = Font(bold=True)

    title = payload.get("title") or REPORT_LABELS.get(report_type, report_type)
    ws.append([f"SpeakBuddi — Báo cáo {title}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    meta = payload.get("meta") or {}
    if meta.get("from_date") and meta.get("to_date"):
        ws.append(["Khoảng thời gian", f"{meta['from_date']} → {meta['to_date']}"])
    if meta.get("plan_name"):
        ws.append(["Gói thanh toán", meta["plan_name"]])
    ws.append([])

    summary = payload.get("summary_rows") or []
    if not summary:
        ws.append([EMPTY_MSG])
    else:
        for label, value in summary:
            ws.append([label, value])

    detail_rows = payload.get("detail_rows") or []
    detail_headers = payload.get("detail_headers") or []
    if detail_headers:
        ws2 = wb.create_sheet("Chi tiết")
        ws2.append(detail_headers)
        for cell in ws2[1]:
            cell.font = bold
        if not detail_rows:
            ws2.append([EMPTY_MSG])
        else:
            for row in detail_rows:
                ws2.append(row)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_pdf(report_type: str, payload: dict[str, Any]) -> bytes:
    ensure_pdf_fonts()
    bio = BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4, leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    base = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PdfTitle",
        parent=base["Title"],
        fontName=FONT_BOLD,
    )
    normal_style = ParagraphStyle(
        "PdfNormal",
        parent=base["Normal"],
        fontName=FONT_REGULAR,
    )
    heading_style = ParagraphStyle(
        "PdfHeading2",
        parent=base["Heading2"],
        fontName=FONT_BOLD,
    )
    story: list[Any] = []

    title = payload.get("title") or REPORT_LABELS.get(report_type, report_type)
    story.append(Paragraph(f"<b>SpeakBuddi — Báo cáo {title}</b>", title_style))
    story.append(Spacer(1, 0.3 * cm))

    meta = payload.get("meta") or {}
    if meta.get("from_date") and meta.get("to_date"):
        story.append(Paragraph(
            f"Khoảng: {meta['from_date']} → {meta['to_date']}",
            normal_style,
        ))
    if meta.get("plan_name"):
        story.append(Paragraph(f"Gói: {meta['plan_name']}", normal_style))
    story.append(Spacer(1, 0.4 * cm))

    summary = payload.get("summary_rows") or []
    if summary:
        summary_data = [["Chỉ số", "Giá trị"]] + [[str(a), str(b)] for a, b in summary]
        tbl = Table(summary_data, colWidths=[8 * cm, 8 * cm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3525cd")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
            ("FONTNAME", (0, 1), (-1, -1), FONT_REGULAR),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(tbl)
    else:
        story.append(Paragraph(EMPTY_MSG, normal_style))

    detail_rows = payload.get("detail_rows") or []
    detail_headers = payload.get("detail_headers") or []
    if detail_headers:
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph("<b>Chi tiết</b>", heading_style))
        if not detail_rows:
            story.append(Paragraph(EMPTY_MSG, normal_style))
        else:
            cap = min(len(detail_rows), 50)
            table_data = [detail_headers] + [list(row) for row in detail_rows[:cap]]
            detail_tbl = Table(table_data, repeatRows=1)
            detail_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eae6f4")),
                ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, -1), FONT_REGULAR),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]))
            story.append(detail_tbl)
            if len(detail_rows) > cap:
                story.append(Paragraph(
                    f"(Hiển thị {cap}/{len(detail_rows)} dòng — xem Excel để đủ dữ liệu)",
                    normal_style,
                ))

    doc.build(story)
    return bio.getvalue()
